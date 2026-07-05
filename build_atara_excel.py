# -*- coding: utf-8 -*-
import pdfplumber, re, glob, os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path(__file__).parent / "עטרי 2025"
FOLDERS = [
    ("קבלות 2025", "קבלות 2025"),
    ("קבלות גלי צאנז 11-25", "גלי צאנז 11-25"),
    ("קבלות גלי צאנז 2 סדנאות ראשונות", "גלי צאנז - סדנאות ראשונות"),
]

HEB = re.compile(r'[֐-׿]')

# נתונים שחולצו ידנית מקבצים סרוקים (תמונה ללא טקסט)
MANUAL = {
    'קבלה 103 רותי אפפלדור.pdf': {
        'num': '103', 'date': '04/11/2025', 'customer': 'רות אפפלדור',
        'total': 650, 'pay_method': 'העברה בנקאית', 'pay_date': '04/11/2025',
        'notes': 'עבור ייעוץ שינה בבית היולדת ב-4.11.25',
    },
    'קבלה מעודכנת הודיה זוהר.pdf': {
        'num': '33', 'date': '03/02/2025', 'customer': 'הודיה זוהר',
        'total': 600, 'pay_method': 'מזומן', 'pay_date': '03/02/2025',
        'notes': 'עבור ייעוץ שינה בבית היולדת ב-3.2.25',
    },
}

def rev_tokens(s):
    """Un-reverse pdfplumber RTL line: reverse token order + reverse hebrew tokens."""
    toks = s.split()[::-1]
    out = []
    for t in toks:
        if HEB.search(t):
            out.append(t[::-1])
        else:
            out.append(t)
    return ' '.join(out)

def money(s):
    m = re.search(r'₪\s*(-?[\d,]+\.?\d*)', s)
    return float(m.group(1).replace(',', '')) if m else None

def parse(pdf_path, group):
    fname = pdf_path.name
    low = fname
    cancelled = any(k in low for k in ['בוטל', 'ביטול', 'מבוטל'])
    with pdfplumber.open(pdf_path) as pdf:
        text = "\n".join((p.extract_text() or "") for p in pdf.pages)
    lines = [l.strip() for l in text.split('\n') if l.strip()]

    rec = {
        'group': group, 'num': None, 'date': None, 'customer': '',
        'total': None, 'tax_deduct': None, 'net': None,
        'pay_method': '', 'pay_details': '', 'pay_date': '',
        'notes': '', 'status': 'מבוטלת' if cancelled else 'פעילה',
        'file': fname,
    }

    if not lines:
        # scanned image, no text - use manual data extracted by eye, else flag
        if fname in MANUAL:
            rec.update(MANUAL[fname])
        else:
            m = re.search(r'(\d{2,3})', fname)
            if m:
                rec['num'] = m.group(1)
            rec['customer'] = '(קובץ סרוק - לבדיקה ידנית)'
            rec['notes'] = 'אין טקסט בקובץ (סריקה/תמונה)'
        return rec

    for l in lines:
        # receipt number: line with קבלה (reversed הלבק)
        if rec['num'] is None and 'הלבק' in l:
            m = re.search(r'(\d{2,4})', l)
            if m:
                rec['num'] = m.group(1)
        # issue date: line with הופק (reversed קפוה)
        if rec['date'] is None and 'קפוה' in l:
            m = re.search(r'(\d{2}/\d{2}/\d{4})', l)
            if m:
                rec['date'] = m.group(1)
        # customer: line with עבור (reversed רובע) and מאת (תאמ)
        if not rec['customer'] and 'רובע' in l:
            fixed = rev_tokens(l)
            m = re.search(r'עבור:\s*(.+)$', fixed)
            if m:
                rec['customer'] = m.group(1).strip()

    # totals
    for l in lines:
        if 'כ"הס' in l and 'ינפל' in l:           # סה"כ לפני ניכוי מס במקור
            rec['net'] = money(l)
        elif 'יוכינ' in l and 'סמ' in l:           # ניכוי מס במקור
            rec['tax_deduct'] = money(l)
        elif ('כ״הס' in l or 'כ"הס' in l):         # סה"כ
            if rec['total'] is None:
                rec['total'] = money(l)

    if rec['total'] is None:
        rec['total'] = rec['net']

    # payment lines: between header and סה"כ
    methods, details, pdates = [], [], []
    in_pay = False
    for l in lines:
        if 'םולשת יעצמא' in l or ('יעצמא' in l and 'םוכס' in l):  # header
            in_pay = True
            continue
        if in_pay:
            if 'כ״הס' in l or 'כ"הס' in l:
                break
            amt = money(l)
            if amt is None:
                continue
            md = re.search(r'(\d{2}/\d{2}/\d{4})', l)
            if md:
                pdates.append(md.group(1))
            if 'ןמוזמ' in l:
                methods.append('מזומן')
            elif "ק'צ" in l or "צ'ק" in l or 'קיש' in l:
                methods.append("צ'ק")
            elif 'הרבעה' in l:
                methods.append('העברה בנקאית')
            elif 'יארשא' in l:
                methods.append('אשראי')
            elif 'טיב' in l or "'יב" in l:
                methods.append('ביט')
            else:
                methods.append('אחר')
            # payment detail = the line minus ₪amount and date, fixed
            d = re.sub(r'₪\s*-?[\d,]+\.?\d*', '', l)
            d = re.sub(r'\d{2}/\d{2}/\d{4}', '', d).strip()
            if d:
                details.append(rev_tokens(d))
    # de-dup methods preserving order
    seen = []
    for mth in methods:
        if mth not in seen:
            seen.append(mth)
    rec['pay_method'] = ' + '.join(seen)
    rec['pay_details'] = ' | '.join([d for d in details if d])
    rec['pay_date'] = ' | '.join(dict.fromkeys(pdates))

    # notes: line with הערות (reversed תורעה)
    for l in lines:
        if 'תורעה' in l:
            fixed = rev_tokens(l)
            m = re.search(r'הערות:\s*(.+)$', fixed)
            rec['notes'] = (m.group(1).strip() if m else fixed.replace('הערות:', '').strip())
            break
    return rec

# ---- collect ----
records = []
for folder, group in FOLDERS:
    fp = BASE / folder
    for pdf in sorted(fp.glob('*.pdf')):
        try:
            records.append(parse(pdf, group))
        except Exception as e:
            print("ERR", pdf.name, e)

def sort_key(r):
    try:
        return int(r['num'])
    except (TypeError, ValueError):
        return 99999
records.sort(key=lambda r: (r['group'], sort_key(r)))

print(f"records: {len(records)}")
print("with total:", sum(1 for r in records if r['total'] is not None))
print("cancelled:", sum(1 for r in records if r['status'] == 'מבוטלת'))

# ---- Excel ----
wb = Workbook()
HDR = PatternFill('solid', fgColor='2E5E4E')
ALT = PatternFill('solid', fgColor='EAF3EE')
WHT = PatternFill('solid', fgColor='FFFFFF')
CANCEL = PatternFill('solid', fgColor='FBE4E4')
TOT = PatternFill('solid', fgColor='D9EAD3')
thin = Side(style='thin', color='C9C9C9')
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)
FN = 'Arial'

def header(ws, headers, widths):
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(name=FN, bold=True, color='FFFFFF', size=11)
        cell.fill = HDR
        cell.alignment = Alignment('center', 'center', wrap_text=True)
        cell.border = BORD
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = 'A2'

ws = wb.active
ws.title = "כל הקבלות"
ws.sheet_view.rightToLeft = True
H = ['#', 'קבוצה', 'מס׳ קבלה', 'תאריך', 'לקוח', 'סכום (₪)',
     'אמצעי תשלום', 'פרטי תשלום', 'תאריך תשלום', 'הערות', 'סטטוס', 'שם קובץ']
W = [5, 20, 9, 12, 26, 12, 16, 30, 13, 34, 9, 30]
header(ws, H, W)
r = 2
for idx, rec in enumerate(records, 1):
    vals = [idx, rec['group'], rec['num'], rec['date'], rec['customer'], rec['total'],
            rec['pay_method'], rec['pay_details'], rec['pay_date'], rec['notes'],
            rec['status'], rec['file']]
    base = CANCEL if rec['status'] == 'מבוטלת' else (ALT if r % 2 == 0 else WHT)
    for c, v in enumerate(vals, 1):
        cell = ws.cell(r, c, v)
        cell.font = Font(name=FN, size=10)
        cell.fill = base
        cell.border = BORD
        if c in (5, 8, 10):
            cell.alignment = Alignment('right', 'center', wrap_text=True)
        else:
            cell.alignment = Alignment('center', 'center', wrap_text=True)
        if c == 6:
            cell.number_format = '#,##0.00'
    r += 1
# totals row (active receipts only)
tr = r
ws.cell(tr, 5, 'סה"כ (קבלות פעילות)').font = Font(name=FN, bold=True, size=11)
for c in range(1, len(H) + 1):
    cell = ws.cell(tr, c)
    cell.fill = TOT
    cell.border = BORD
    cell.alignment = Alignment('center', 'center')
ws.cell(tr, 5).alignment = Alignment('right', 'center')
active_sum = sum(r['total'] for r in records if r['status'] == 'פעילה' and r['total'])
c6 = ws.cell(tr, 6, round(active_sum, 2))
c6.font = Font(name=FN, bold=True, size=11)
c6.number_format = '#,##0.00'
c6.fill = TOT
c6.border = BORD

# ---- summary sheet ----
ws2 = wb.create_sheet("סיכום לפי קבוצה")
ws2.sheet_view.rightToLeft = True
H2 = ['קבוצה', 'מס׳ קבלות פעילות', 'מס׳ קבלות מבוטלות', 'סה"כ הכנסות פעילות (₪)']
W2 = [28, 18, 18, 24]
header(ws2, H2, W2)
r = 2
groups = []
for _, g in FOLDERS:
    if g not in groups:
        groups.append(g)
grand = 0
for g in groups:
    act = [x for x in records if x['group'] == g and x['status'] == 'פעילה']
    can = [x for x in records if x['group'] == g and x['status'] == 'מבוטלת']
    s = sum(x['total'] for x in act if x['total'])
    grand += s
    vals = [g, len(act), len(can), round(s, 2)]
    for c, v in enumerate(vals, 1):
        cell = ws2.cell(r, c, v)
        cell.font = Font(name=FN, size=10)
        cell.fill = ALT if r % 2 == 0 else WHT
        cell.border = BORD
        cell.alignment = Alignment('center', 'center')
        if c == 4:
            cell.number_format = '#,##0.00'
    r += 1
# grand total
for c, v in enumerate(['סה"כ כללי', '', '', round(grand, 2)], 1):
    cell = ws2.cell(r, c, v)
    cell.font = Font(name=FN, bold=True, size=11)
    cell.fill = TOT
    cell.border = BORD
    cell.alignment = Alignment('center', 'center')
    if c == 4:
        cell.number_format = '#,##0.00'

out = str(Path(__file__).parent / "קבלות_עטרי_2025_מפורט.xlsx")
wb.save(out)
print("saved:", out)
