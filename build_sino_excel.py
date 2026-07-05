import os
import re
import sys
from datetime import datetime

# Ensure UTF-8 output on Windows
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Installing openpyxl...")
    os.system("pip install openpyxl")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SOURCE_DIR = r"C:\Users\בונים ומוגנים\OneDrive\שולחן העבודה\קלוד\סינו"
OUTPUT_FILE = os.path.join(SOURCE_DIR, "סינומקס_מסמכים.xlsx")

# ---- Category rules (order matters — first match wins) ----
CATEGORIES = [
    ("חשבוניות מס", "E8F5E9", lambda n: (
        re.search(r'Tax_Invoice', n, re.IGNORECASE) or
        re.search(r'HashDoc', n, re.IGNORECASE) or
        re.search(r'חשבונית מס', n) or
        re.search(r'חשבונית_מס', n) or
        re.search(r'^(50|60|61|70)\d+', n) or
        re.search(r'חשבונית [56]', n) or
        re.search(r'^סינומקס', n)
    )),
    ("קבלות", "E3F2FD", lambda n: (
        re.search(r'receipt', n, re.IGNORECASE) or
        re.search(r'קבלה', n)
    )),
    ("פרופורמות", "FFF9C4", lambda n: re.search(r'Proforma', n, re.IGNORECASE)),
    ("מס הקצאה", "FCE4EC", lambda n: (
        re.search(r'מס הקצאה', n) or
        re.search(r'הקצאה', n) or
        re.search(r'50594 מס הקצאה', n)
    )),
    ("כרטסות", "EDE7F6", lambda n: re.search(r'כרטסת', n)),
    ("תמונות וצ'קים", "FFF3E0", lambda n: (
        re.search(r'WhatsApp Image', n, re.IGNORECASE) or
        re.search(r'^IMG_', n, re.IGNORECASE) or
        re.search(r'^IMG-', n, re.IGNORECASE)
    )),
    ("מסמכים סרוקים", "E0F7FA", lambda n: (
        re.search(r'CamScanner', n, re.IGNORECASE) or
        re.search(r'Scanned', n, re.IGNORECASE) or
        re.search(r'WhatsApp Scan', n, re.IGNORECASE)
    )),
    ("קבצי אקסל", "F3E5F5", lambda n: re.search(r'\.(xlsx|csv)$', n, re.IGNORECASE)),
]

def categorize(filename):
    for cat_name, color, test in CATEGORIES:
        if test(filename):
            return cat_name, color
    return "אחר", "F5F5F5"

def extract_doc_number(filename):
    """Extract the first sequence of 4+ digits from the filename."""
    m = re.search(r'\d{4,}', filename)
    return m.group(0) if m else ""

def get_file_date(filepath):
    try:
        ts = os.path.getmtime(filepath)
        return datetime.fromtimestamp(ts).strftime("%d/%m/%Y")
    except Exception:
        return ""

def hex_fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def header_font():
    return Font(bold=True, size=11)

def thin_border():
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)

def write_sheet(ws, rows, header_color):
    headers = ["#", "שם קובץ", "מספר מסמך", "תאריך קובץ", "קטגוריה", "הערות"]
    col_widths = [5, 55, 16, 14, 18, 20]

    # Write headers
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = hex_fill(header_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # Write data rows
    for row_idx, row_data in enumerate(rows, start=2):
        alt_fill = hex_fill("FAFAFA") if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="right" if col_idx > 1 else "center",
                                       vertical="center")
            if alt_fill:
                cell.fill = alt_fill
        ws.row_dimensions[row_idx].height = 18

# ---- Main ----
all_files = sorted(os.listdir(SOURCE_DIR))
# Skip the output file itself and directories
all_files = [f for f in all_files
             if os.path.isfile(os.path.join(SOURCE_DIR, f))
             and f != "סינומקס_מסמכים.xlsx"]

# Categorize everything
categorized = {}  # cat_name -> list of (filename, doc_num, date, category, notes)
for fname in all_files:
    cat_name, color = categorize(fname)
    fpath = os.path.join(SOURCE_DIR, fname)
    doc_num = extract_doc_number(fname)
    fdate = get_file_date(fpath)
    if cat_name not in categorized:
        categorized[cat_name] = {"color": color, "rows": []}
    categorized[cat_name]["rows"].append((fname, doc_num, fdate, cat_name, ""))

# Define sheet order
SHEET_ORDER = [
    "חשבוניות מס",
    "קבלות",
    "פרופורמות",
    "מס הקצאה",
    "כרטסות",
    "תמונות וצ'קים",
    "מסמכים סרוקים",
    "קבצי אקסל",
    "אחר",
]

# Header colors per sheet (dark enough for white text)
HEADER_COLORS = {
    "חשבוניות מס":    "2E7D32",
    "קבלות":          "1565C0",
    "פרופורמות":      "F9A825",
    "מס הקצאה":       "C62828",
    "כרטסות":         "6A1B9A",
    "תמונות וצ'קים":  "E65100",
    "מסמכים סרוקים":  "00838F",
    "קבצי אקסל":      "6A1B9A",
    "אחר":            "546E7A",
}

wb = openpyxl.Workbook()
wb.remove(wb.active)  # Remove default sheet

# ---- Summary sheet ----
ws_summary = wb.create_sheet("סיכום")
ws_summary.sheet_view.rightToLeft = True

summary_headers = ["קטגוריה", "כמות קבצים", "צבע"]
summary_col_widths = [25, 15, 12]

for col_idx, (h, w) in enumerate(zip(summary_headers, summary_col_widths), start=1):
    cell = ws_summary.cell(row=1, column=col_idx, value=h)
    cell.font = Font(bold=True, size=12, color="FFFFFF")
    cell.fill = hex_fill("37474F")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border()
    ws_summary.column_dimensions[cell.column_letter].width = w

ws_summary.row_dimensions[1].height = 24
ws_summary.freeze_panes = "A2"

total_files = 0
for row_idx, cat_name in enumerate(SHEET_ORDER, start=2):
    data = categorized.get(cat_name, {"rows": [], "color": "EEEEEE"})
    count = len(data["rows"])
    total_files += count
    hdr_color = HEADER_COLORS.get(cat_name, "546E7A")

    cat_cell = ws_summary.cell(row=row_idx, column=1, value=cat_name)
    cat_cell.font = Font(bold=True, size=11)
    cat_cell.fill = hex_fill(data["color"])
    cat_cell.alignment = Alignment(horizontal="right", vertical="center")
    cat_cell.border = thin_border()

    count_cell = ws_summary.cell(row=row_idx, column=2, value=count)
    count_cell.font = Font(size=11)
    count_cell.fill = hex_fill(data["color"])
    count_cell.alignment = Alignment(horizontal="center", vertical="center")
    count_cell.border = thin_border()

    color_cell = ws_summary.cell(row=row_idx, column=3, value="")
    color_cell.fill = hex_fill(hdr_color)
    color_cell.border = thin_border()

    ws_summary.row_dimensions[row_idx].height = 20

# Total row
total_row = len(SHEET_ORDER) + 2
for col_idx, val in enumerate(["סה\"כ", total_files, ""], start=1):
    cell = ws_summary.cell(row=total_row, column=col_idx, value=val)
    cell.font = Font(bold=True, size=12, color="FFFFFF")
    cell.fill = hex_fill("37474F")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border()
ws_summary.row_dimensions[total_row].height = 22

# ---- Category sheets ----
for cat_name in SHEET_ORDER:
    data = categorized.get(cat_name, {"rows": [], "color": "EEEEEE"})
    rows = data["rows"]
    hdr_color = HEADER_COLORS.get(cat_name, "546E7A")

    ws = wb.create_sheet(cat_name)
    ws.sheet_view.rightToLeft = True

    numbered_rows = [(i + 1, fname, doc_num, fdate, category, notes)
                     for i, (fname, doc_num, fdate, category, notes) in enumerate(rows)]
    write_sheet(ws, numbered_rows, hdr_color)

    if not rows:
        ws.cell(row=2, column=2, value="אין קבצים בקטגוריה זו")

wb.save(OUTPUT_FILE)
print(f"Done. {total_files} files processed.")
print(f"Saved: {OUTPUT_FILE}")

# Print summary
for cat_name in SHEET_ORDER:
    count = len(categorized.get(cat_name, {}).get("rows", []))
    if count:
        print(f"  {cat_name}: {count}")
