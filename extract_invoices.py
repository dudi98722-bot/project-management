import pdfplumber
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent / "ירוקה"

HEB = re.compile(r'[֐-׿]')

def fix_heb(s):
    """Un-reverse pdfplumber RTL output: reverse token order, reverse each Hebrew token."""
    toks = s.split()
    toks = toks[::-1]
    out = []
    for t in toks:
        if HEB.search(t) and not re.search(r'\d', t):
            out.append(t[::-1])
        elif HEB.search(t):
            # mixed hebrew+digit: reverse only, best effort
            out.append(t[::-1])
        else:
            out.append(t)
    return ' '.join(out)

def money(s):
    """Return first ₪ amount as float from a line."""
    m = re.search(r'₪\s*([\d,]+\.?\d*)', s)
    if m:
        return float(m.group(1).replace(',', ''))
    return None

def all_money(s):
    return [float(x.replace(',', '')) for x in re.findall(r'₪\s*([\d,]+\.?\d*)', s)]

def extract(pdf_path, inv_type):
    with pdfplumber.open(pdf_path) as pdf:
        text = "\n".join((p.extract_text() or "") for p in pdf.pages)
    lines = [l for l in text.split('\n') if l.strip()]

    inv_num = None
    date = None
    customer = None
    vat = None
    vat_pct = None
    total = None
    subtotal = None
    payment_method = None
    ref = None

    # date = first dd/mm/yyyy
    for l in lines:
        m = re.search(r'\b(\d{2}/\d{2}/\d{4})\b', l)
        if m:
            date = m.group(1)
            break

    # invoice number from filename (reliable): 2024.50005 -> 50005
    fn = pdf_path.stem
    if '.' in fn:
        inv_num = fn.split('.')[-1]

    # customer = line index 1 (reversed), unless it's the doc-title line
    if len(lines) > 1:
        cand = lines[1]
        if 'תינובשח' not in cand and 'כ"הס' not in cand:
            customer = fix_heb(cand)

    for l in lines:
        # VAT line: contains reversed מע"מ = מ"עמ  and a percent
        if 'מ"עמ' in l or re.search(r'\d+%', l):
            if 'מ"עמ' in l:
                vat = money(l)
                pm = re.search(r'(\d+)%', l)
                if pm:
                    vat_pct = int(pm.group(1))
        # total: line with לתשלום (reversed םולשתל) and סה"כ (כ"הס)
        if 'םולשתל' in l and 'כ"הס' in l:
            total = money(l)

    if total is not None and vat is not None:
        subtotal = round(total - vat, 2)

    # payment method (חשבונית מס קבלה)
    for l in lines:
        if 'הרבעה' in l:
            payment_method = 'העברה בנקאית'
        elif 'ןמוזמ' in l:
            payment_method = 'מזומן'
        elif "ק'צ" in l or 'קיש' in l or "צ'ק" in l:
            payment_method = "צ'ק / שיק"
        elif 'יארשא' in l:
            payment_method = 'אשראי'
        elif "'יב" in l or 'יב יד' in l or "ביט" in l:
            payment_method = 'ביט'
    # reference (עבור חשבון עסקה NNNNN)
    for l in lines:
        if 'הקסע' in l:
            m = re.search(r'(\d{4,})', l)
            if m:
                ref = m.group(1)

    # items: between header (has תומכ & ריחמ) and first totals line
    items = []
    has_makat = any('ט"קמ' in l for l in lines)
    in_items = False
    for l in lines:
        if ('תומכ' in l and 'ריחמ' in l):
            in_items = True
            continue
        if in_items:
            # stop at totals/notes
            if ('כ"הס' in l) or ('מ"עמ' in l) or ('לוגיע' in l) or ('יטרפ' in l) or ('המיתח' in l):
                in_items = False
                continue
            amts = all_money(l)
            if len(amts) >= 2:
                line_total = amts[0]
                unit_price = amts[1]
                rest = re.sub(r'₪\s*[\d,]+\.?\d*', '', l).strip()
                toks = rest.split()
                makat = None
                qty = None
                if has_makat and toks:
                    makat = toks[-1]
                    toks = toks[:-1]
                if toks:
                    qty = toks[-1]
                    desc_toks = toks[:-1]
                else:
                    desc_toks = []
                desc = fix_heb(' '.join(desc_toks))
                items.append({
                    'makat': makat or '',
                    'qty': qty or '',
                    'desc': desc,
                    'unit': unit_price,
                    'total': line_total,
                })
    return {
        'type': inv_type,
        'num': inv_num,
        'date': date,
        'customer': customer or '',
        'subtotal': subtotal,
        'vat': vat,
        'vat_pct': vat_pct,
        'total': total,
        'payment': payment_method or '',
        'ref': ref or '',
        'items': items,
        'file': pdf_path.name,
    }

invoices = []
for folder, label in [('חשבונית מס', 'חשבונית מס'), ('חשבונית מס קבלה', 'חשבונית מס/קבלה')]:
    fp = BASE_DIR / folder
    if not fp.exists():
        continue
    for pdf in sorted(fp.glob('*.pdf')):
        try:
            invoices.append(extract(pdf, label))
        except Exception as e:
            print(f"ERR {pdf.name}: {e}")

invoices.sort(key=lambda x: (x['date'] and tuple(reversed(x['date'].split('/'))) or '', x['num'] or ''))

print(f"invoices: {len(invoices)}")
ok_cust = sum(1 for i in invoices if i['customer'])
ok_total = sum(1 for i in invoices if i['total'] is not None)
print(f"with customer: {ok_cust} | with total: {ok_total}")

# ---------- Build Excel ----------
wb = Workbook()
HDR = PatternFill('solid', fgColor='1F4E5F')
ALT = PatternFill('solid', fgColor='EAF3F6')
WHT = PatternFill('solid', fgColor='FFFFFF')
TOT = PatternFill('solid', fgColor='D9EAD3')
thin = Side(style='thin', color='C9C9C9')
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)
FN = 'Arial'

def style_header(ws, headers, widths):
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(name=FN, bold=True, color='FFFFFF', size=11)
        cell.fill = HDR
        cell.alignment = Alignment('center', 'center', wrap_text=True)
        cell.border = BORD
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'

# Sheet 1: summary
ws = wb.active
ws.title = "סיכום חשבוניות"
ws.sheet_view.rightToLeft = True
H1 = ['#', 'סוג מסמך', 'מס׳ חשבונית', 'תאריך', 'לקוח', 'סה"כ לפני מע"מ (₪)', 'מע"מ (₪)', 'אחוז מע"מ', 'סה"כ לתשלום (₪)', 'אמצעי תשלום', 'אסמכתא/עסקה', 'קובץ']
W1 = [5, 16, 12, 12, 34, 17, 13, 9, 17, 16, 13, 20]
style_header(ws, H1, W1)
r = 2
for idx, inv in enumerate(invoices, 1):
    vals = [idx, inv['type'], inv['num'], inv['date'], inv['customer'], inv['subtotal'],
            inv['vat'], (f"{inv['vat_pct']}%" if inv['vat_pct'] else ''), inv['total'],
            inv['payment'], inv['ref'], inv['file']]
    fill = ALT if r % 2 == 0 else WHT
    for c, v in enumerate(vals, 1):
        cell = ws.cell(r, c, v)
        cell.font = Font(name=FN, size=10)
        cell.fill = fill
        cell.border = BORD
        if c == 5:
            cell.alignment = Alignment('right', 'center')
        else:
            cell.alignment = Alignment('center', 'center')
        if c in (6, 7, 9):
            cell.number_format = '#,##0.00'
    r += 1
# totals
tr = r
ws.cell(tr, 5, 'סה"כ').font = Font(name=FN, bold=True, size=11)
for c in range(1, len(H1) + 1):
    cell = ws.cell(tr, c)
    cell.fill = TOT
    cell.border = BORD
    cell.alignment = Alignment('center', 'center')
ws.cell(tr, 5).fill = TOT
for c, col in [(6, 'F'), (7, 'G'), (9, 'I')]:
    cell = ws.cell(tr, c, f'=SUM({col}2:{col}{tr-1})')
    cell.font = Font(name=FN, bold=True, size=11)
    cell.number_format = '#,##0.00'
    cell.fill = TOT
    cell.border = BORD

# Sheet 2: items
ws2 = wb.create_sheet("פירוט שורות")
ws2.sheet_view.rightToLeft = True
H2 = ['מס׳ חשבונית', 'תאריך', 'לקוח', 'סוג מסמך', 'מק"ט', 'כמות', 'פירוט', 'מחיר יחידה (₪)', 'סה"כ שורה (₪)']
W2 = [12, 12, 30, 16, 12, 8, 42, 15, 15]
style_header(ws2, H2, W2)
r = 2
for inv in invoices:
    for it in inv['items']:
        vals = [inv['num'], inv['date'], inv['customer'], inv['type'], it['makat'],
                it['qty'], it['desc'], it['unit'], it['total']]
        fill = ALT if r % 2 == 0 else WHT
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(r, c, v)
            cell.font = Font(name=FN, size=10)
            cell.fill = fill
            cell.border = BORD
            if c in (3, 7):
                cell.alignment = Alignment('right', 'center')
            else:
                cell.alignment = Alignment('center', 'center')
            if c in (8, 9):
                cell.number_format = '#,##0.00'
        r += 1

out = str(BASE_DIR / "חשבוניות_ירוקה_מפורט.xlsx")
wb.save(out)
print("saved:", out)
