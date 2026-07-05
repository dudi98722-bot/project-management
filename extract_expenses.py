# -*- coding: utf-8 -*-
"""חילוץ מפורט של תיקיית 'הוצאות 25/קבלות למס' לאקסל."""
import pdfplumber, re, glob, os
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path(__file__).parent / "הוצאות 25" / "קבלות למס"
HEB = re.compile(r'[֐-׿]')

def fix_heb(s):
    toks = s.split()[::-1]
    return ' '.join(t[::-1] if HEB.search(t) else t for t in toks)

def nums(s):
    res = []
    for m in re.finditer(r'\(?-?[\d,]+\.?\d*\)?', s):
        tok = m.group(0)
        if not re.search(r'\d', tok):
            continue
        neg = tok.startswith('(') and tok.endswith(')')
        tok = tok.strip('()').replace(',', '')
        try:
            v = float(tok)
        except ValueError:
            continue
        res.append(-v if neg else v)
    return res

VENDOR_BY_ID = {
    '515269488': 'קול כשר בע"מ',
    '515984094': 'מטרה הפקות בע"מ',
    '204673982': 'קומפיוטר - רשת חדרי מחשבים',
    '514204775': 'QONLINE DIR',
    '513797506': 'מפתח העיר 2006 בע"מ',
    '515764306': 'חזק מגזין פרסום בע"מ',
    '036589026': 'רחל רוטנברג',
    '589903046': 'החברה לרפואת הנקה בישראל',
    '580470367': 'האיגוד הישראלי למקצועות ההנקה',
    '580066421': 'ליגת לה לצ\'ה ישראל',
    '514602879': 'דפוס רימון בע"מ',
    '325027423': 'שאר ר ציפורה (BAZ) - עיצוב גרפי',
    '512711789': 'גמא ניהול וסליקה בע"מ',
}

# עקיפות לפי שם קובץ (מקרים שקשה לזהות אוטומטית)
FNAME_OVERRIDE = {
    'InvoiceReceipt_68817.pdf':  dict(vendor='האיגוד הישראלי למקצועות ההנקה', desc='כנס האיגוד הישראלי למקצועות ההנקה 4.2.25'),
    '871243.pdf':                dict(vendor='תרומה - טהרת המשפחה', desc='תרומה (סעיף 46)', subtotal=None, vat=None, total=180.0),
    'ע..פלדמן  1250.pdf':         dict(desc='לוגו ושפה עיצובית, תפילה, טפסים, סילבוס ועוד'),
    'עטרה פלדמן 295.pdf':         dict(desc='הדפסות כרומו 300 גרם'),
    'עטרה פלדמן.pdf':            dict(desc='פרסום במדריך'),
}

# ספקי קבצי cid (טקסט לא קריא) – זוהו ידנית כמנוי חודשי 88.50
CID_VENDOR = 'מנוי חודשי (טקסט מוצפן - לאימות)'

def is_gama_deposit(fname):
    return 'AUTODIGIT' in fname.upper() or fname.lower().startswith('fax-')

NOISE_AMT = ['במחשב', 'עסקה', 'שובר', 'פרעון', 'אישור', 'ref', 'כרטיס',
             'ספרות', 'ברוטו', 'פעימות', 'שוברים']
DESC_NOISE = ['כרטיס אשראי', 'אופן חיוב', 'מספר תשלומים', 'סכום עסקה', 'עסקה מספר',
              'שובר', 'פרעון', 'שע"ח', "שע''ח", 'שקל חדש', 'מחשבונית', 'סוג עסקה',
              'ספרות', 'תשלום באמצעות', 'חיוב אוטומטי', 'חיוב מהטלפון', 'שורהפרטים',
              'מספר אישור', 'קבלה מחשבונית', 'Total', 'Payment', 'Credit', 'ref.',
              'ILS', 'במחשב', 'תאריך', 'לתשלום', 'סכום פר תשלום', 'הננו לעדכנך',
              'סכום כל', 'עמלות', 'הנחות', 'התקבולים', 'ריבית', 'אשראי', 'ביניים',
              'עטרה פלדמן', 'ערך צבור', 'דקות:', 'מזומן']

def pick(fixed_lines, labels, avoid=()):
    cands = []
    for fl in fixed_lines:
        if any(n in fl for n in NOISE_AMT):
            continue
        if any(a in fl for a in avoid):
            continue
        if any(lb in fl for lb in labels):
            for n in nums(fl):
                if 1900 <= n <= 2100 and n == int(n):
                    continue
                if n in (18.0, 17.0):
                    continue
                cands.append(n)
    return cands

def clean_desc(fixed_lines):
    out = []
    for fl in fixed_lines:
        if not HEB.search(fl):
            continue
        if any(sw in fl for sw in (['סה"כ', 'מע"מ', 'הנחה', 'חתימה', 'הופק', 'בנק',
                                    'מפיק', 'מנפיק', 'דובכל', 'לכבוד', 'פטור'] + DESC_NOISE)):
            continue
        if re.search(r'\d+\.\d{2}|₪', fl) and len(re.sub(r'[\d.,%/:\-()₪"\'‏]', '', fl).strip()) > 6:
            txt = re.sub(r'[\d.,%₪/]+', ' ', fl)
            txt = re.sub(r'\s+', ' ', txt).strip(' #:-|')
            if len(txt) > 5 and txt not in out:
                out.append(txt)
        if len(out) >= 2:
            break
    return ' | '.join(out)[:160]

def find_vendor(raw_lines, fixed_lines, fname):
    blob = "\n".join(raw_lines).replace('-', '')
    for vid, name in VENDOR_BY_ID.items():
        if vid in blob:
            return name
    if is_gama_deposit(fname):
        return 'גמא ניהול וסליקה בע"מ'
    skip = ('לכבוד', 'דובכל', 'בס"ד', 'ד"סב', 'מקור', 'רוקמ', 'חשבונית', 'קבלה',
            'מספר', 'דו"ח', 'הופק', 'דף', 'תאריך', 'ידידינו', 'הננו')
    for fl in fixed_lines:
        if 'מפיק המסמך' in fl or fl.startswith('מנפיק'):
            v = re.sub(r'.*מפיק[ה]? ?(המסמך)?\s*:?', '', fl).strip(' :[]')
            if v and HEB.search(v):
                return v
    for fl in fixed_lines:
        f = fl.strip()
        if len(f) < 4 or not HEB.search(f):
            continue
        if any(f.startswith(s) or s in f[:6] for s in skip):
            continue
        return f
    return ''

def extract(path):
    with pdfplumber.open(path) as pdf:
        raw = "\n".join((p.extract_text() or "") for p in pdf.pages)
        meta = pdf.metadata or {}
    raw_lines = [l for l in raw.split('\n') if l.strip()]
    fixed = [fix_heb(l) for l in raw_lines]
    is_cid = '(cid:' in raw
    fname = path.name
    is_gama_dep = is_gama_deposit(fname)
    is_gama_comm = ('512711789' in raw.replace('-', '')) and not is_gama_dep

    vendor = find_vendor(raw_lines, fixed, fname)
    if is_cid:
        vendor = CID_VENDOR

    # תאריך
    date = ''
    for l in raw_lines:
        m = re.search(r'\b(\d{2})[/.\-](\d{2})[/.\-](\d{4}|\d{2})\b', l)
        if m:
            d, mo, y = m.groups()
            y = '20' + y if len(y) == 2 else y
            if 2024 <= int(y) <= 2027:
                date = f'{d}/{mo}/{y}'; break

    # מספר מסמך
    docnum = ''
    for fl in fixed:
        m = re.search(r'(?:חשבונית|קבלה|מסמך|Receipt)[^\d]*?(?:מספר|:)?[^\d]*?([\d/]{3,})', fl)
        if m:
            docnum = m.group(1).strip('/'); break
    if not docnum:
        m = re.search(r'(\d{3,})', path.stem)
        docnum = m.group(1) if m else ''

    subtotal = vat = total = None
    if is_gama_comm:
        c = pick(fixed, ['כולל מע"מ'])
        total = min(c) if c else None          # עמלה כולל מע"מ
        c = pick(fixed, ['מע"מ 18', '18% מע"מ', '18.00%'])
        vat = min(c) if c else None
        if total is not None and vat is not None:
            subtotal = round(total - vat, 2)
        doctype = 'חשבונית עמלת סליקה (גמא)'
    else:
        if is_cid:
            # רק אסימונים עם 2 ספרות אחרי הנקודה (סכומי כסף אמיתיים), ללא 0.00 ו-18.00
            twodec = [float(m.group(0).replace(',', ''))
                      for m in re.finditer(r'\d[\d,]*\.\d{2}\b', raw)]
            twodec = [n for n in twodec if 1 < n < 100000 and n != 18.0]
            total = max(twodec) if twodec else None
            for l in raw_lines:
                if '18.00%' in l or '18%' in l:
                    vv = [float(m.group(0)) for m in re.finditer(r'\d+\.\d{2}\b', l)
                          if float(m.group(0)) not in (18.0, 0.0) and (total is None or float(m.group(0)) < total)]
                    if vv:
                        vat = max(vv)
            if total is not None and vat is not None:
                subtotal = round(total - vat, 2)
            doctype = 'קבלה (טקסט מוצפן)'
        else:
            c = pick(fixed, ['לתשלום']) or pick(fixed, ['כולל מע"מ'], avoid=['ללא', 'לפני'])
            if not c:
                c = pick(fixed, ['סה"כ', 'סך', 'שולם'],
                         avoid=['לפני', 'ללא', 'ביניים', 'חייב', 'פטור'])
            total = max(c) if c else None
            sc = pick(fixed, ['לפני מע"מ', 'ללא מע"מ', 'חייב במע"מ', 'פטור ממע"מ', 'ביניים'])
            subtotal = max(sc) if sc else None
            for fl in fixed:
                if 'מע"מ' in fl and re.search(r'1[78](\.0+)?\s*%|%\s*1[78]', fl):
                    vv = [n for n in nums(fl) if n not in (18.0, 17.0)
                          and (total is None or n < total) and n != subtotal]
                    if vv:
                        vat = max(vv)
            if vat is None and total and subtotal:
                d2 = round(total - subtotal, 2)
                vat = d2 if d2 > 0 else None
            if subtotal is None and total and vat:
                subtotal = round(total - vat, 2)
            blob = " ".join(fixed)
            doctype = ('חשבונית מס/קבלה' if 'חשבונית מס קבלה' in blob else
                       'חשבונית מס' if 'חשבונית מס' in blob else 'קבלה')

    description = clean_desc(fixed)

    row = dict(vendor=vendor, doctype=doctype, docnum=docnum, date=date,
               subtotal=subtotal, vat=vat, total=total, desc=description,
               file=fname, gama_dep=is_gama_dep,
               flag='בדיקה ידנית' if (is_cid or total is None) else '')

    ov = FNAME_OVERRIDE.get(fname)
    if ov:
        row.update(ov)
        if row.get('total') is not None:
            row['flag'] = ''
    return row

# גמא פיקדון: חילוץ ברוטו + עמלה לגיליון אסמכתאות
def extract_gama_dep(path):
    with pdfplumber.open(path) as pdf:
        raw = "\n".join((p.extract_text() or "") for p in pdf.pages)
    fixed = [fix_heb(l) for l in raw.split('\n') if l.strip()]
    date = ''
    for l in raw.split('\n'):
        m = re.search(r'\b(\d{2})[/.\-](\d{2})[/.\-](\d{4}|\d{2})\b', l)
        if m:
            d, mo, y = m.groups(); y = '20'+y if len(y) == 2 else y
            if 2024 <= int(y) <= 2027:
                date = f'{d}/{mo}/{y}'; break
    gross = fee = None
    g = [n for fl in fixed if 'ברוטו' in fl or 'שוברי כרטיסי' in fl for n in nums(fl) if n > 5]
    if g: gross = max(g)
    f = [n for fl in fixed if 'עמלות סליקה' in fl or 'עמלות תפעוליות' in fl for n in nums(fl)]
    fee = sum(f) if f else None
    m = re.search(r'(\d{3,})', path.stem)
    return dict(date=date, gross=gross, fee=fee, file=path.name,
                ref=(re.search(r'(\d{6,})', path.stem).group(1) if re.search(r'(\d{6,})', path.stem) else ''))

# ---------- איסוף ----------
rows, gama_dep = [], []
for f in sorted(glob.glob(str(BASE / '*.pdf'))):
    p = Path(f)
    try:
        if is_gama_deposit(p.name):
            gama_dep.append(extract_gama_dep(p))
        else:
            rows.append(extract(p))
    except Exception as e:
        rows.append(dict(vendor='', doctype='', docnum='', date='', subtotal=None,
                         vat=None, total=None, desc='', file=p.name, gama_dep=False,
                         flag=f'שגיאה: {e}'))
for f in sorted(glob.glob(str(BASE / '*.jpg')) + glob.glob(str(BASE / '*.jpeg')) + glob.glob(str(BASE / '*.png'))):
    rows.append(dict(vendor='', doctype='תמונה (סריקה)', docnum='', date='', subtotal=None,
                     vat=None, total=None, desc='', file=os.path.basename(f),
                     gama_dep=False, flag='תמונה - בדיקה ידנית'))

def sk(r):
    if r['date']:
        try:
            d, m, y = r['date'].split('/'); return (0, int(y), int(m), int(d))
        except Exception:
            pass
    return (1, 0, 0, 0)
rows.sort(key=sk)
gama_dep.sort(key=lambda r: sk(r))

# סימון כפילויות (אותו ספק+תאריך+סכום, למשל חשבונית+קבלה נפרדות)
seen = {}
for row in rows:
    if row['total'] is None or not row['date']:
        continue
    key = (row['vendor'], row['date'], round(row['total'], 2))
    if key in seen:
        if not row['flag']:
            row['flag'] = 'כפילות? (אותו ספק/תאריך/סכום)'
    else:
        seen[key] = True

# ---------- אקסל ----------
wb = Workbook()
HDR = PatternFill('solid', fgColor='1F4E5F'); ALT = PatternFill('solid', fgColor='EAF3F6')
WHT = PatternFill('solid', fgColor='FFFFFF'); TOT = PatternFill('solid', fgColor='D9EAD3')
FLG = PatternFill('solid', fgColor='FCE8D6'); GAMA = PatternFill('solid', fgColor='F4F4F4')
thin = Side(style='thin', color='C9C9C9'); BORD = Border(thin, thin, thin, thin)
FN = 'Arial'

def header(ws, H, W):
    for c, (h, w) in enumerate(zip(H, W), 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(name=FN, bold=True, color='FFFFFF', size=11)
        cell.fill = HDR; cell.alignment = Alignment('center', 'center', wrap_text=True)
        cell.border = BORD; ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 32; ws.freeze_panes = 'A2'

# גיליון 1 – הוצאות
ws = wb.active; ws.title = "הוצאות 2025"; ws.sheet_view.rightToLeft = True
H = ['#', 'תאריך', 'ספק', 'סוג מסמך', 'מס׳ מסמך', 'תיאור',
     'לפני מע"מ (₪)', 'מע"מ (₪)', 'סה"כ לתשלום (₪)', 'הערה', 'קובץ']
W = [4, 11, 30, 22, 13, 44, 13, 11, 15, 13, 26]
header(ws, H, W)
r = 2
for i, row in enumerate(rows, 1):
    vals = [i, row['date'], row['vendor'], row['doctype'], row['docnum'], row['desc'],
            row['subtotal'], row['vat'], row['total'], row['flag'], row['file']]
    base = ALT if r % 2 == 0 else WHT
    for c, v in enumerate(vals, 1):
        cell = ws.cell(r, c, v)
        cell.font = Font(name=FN, size=10)
        cell.fill = FLG if (row['flag'] and c == 10) else base
        cell.border = BORD
        cell.alignment = Alignment('right' if c in (3, 6) else 'center', 'center',
                                   wrap_text=(c in (3, 6)))
        if c in (7, 8, 9):
            cell.number_format = '#,##0.00'
    r += 1
tr = r
for c in range(1, len(H) + 1):
    ws.cell(tr, c).fill = TOT; ws.cell(tr, c).border = BORD
ws.cell(tr, 6, 'סה"כ הוצאות'); ws.cell(tr, 6).font = Font(name=FN, bold=True, size=11)
ws.cell(tr, 6).alignment = Alignment('right', 'center')
for c, col in [(7, 'G'), (8, 'H'), (9, 'I')]:
    cell = ws.cell(tr, c, f'=SUM({col}2:{col}{tr-1})')
    cell.font = Font(name=FN, bold=True, size=11); cell.number_format = '#,##0.00'
    cell.fill = TOT; cell.border = BORD; cell.alignment = Alignment('center', 'center')

# גיליון 2 – סיכום לפי ספק
ws2 = wb.create_sheet("סיכום לפי ספק"); ws2.sheet_view.rightToLeft = True
agg = defaultdict(lambda: [0, 0.0])
for row in rows:
    k = row['vendor'] or '(לא זוהה)'
    agg[k][0] += 1; agg[k][1] += (row['total'] or 0)
header(ws2, ['ספק', 'מס׳ מסמכים', 'סה"כ (₪)'], [34, 14, 16])
rr = 2
for vend, (cnt, tot) in sorted(agg.items(), key=lambda x: -x[1][1]):
    for c, v in enumerate([vend, cnt, round(tot, 2)], 1):
        cell = ws2.cell(rr, c, v); cell.font = Font(name=FN, size=10)
        cell.fill = ALT if rr % 2 == 0 else WHT; cell.border = BORD
        cell.alignment = Alignment('right' if c == 1 else 'center', 'center')
        if c == 3: cell.number_format = '#,##0.00'
    rr += 1
for c in range(1, 4):
    ws2.cell(rr, c).fill = TOT; ws2.cell(rr, c).border = BORD
ws2.cell(rr, 1, 'סה"כ'); ws2.cell(rr, 1).font = Font(name=FN, bold=True)
ws2.cell(rr, 1).alignment = Alignment('right', 'center')
ws2.cell(rr, 3, f'=SUM(C2:C{rr-1})'); ws2.cell(rr, 3).font = Font(name=FN, bold=True)
ws2.cell(rr, 3).number_format = '#,##0.00'; ws2.cell(rr, 3).alignment = Alignment('center', 'center')

# גיליון 3 – דוחות סליקה גמא (אסמכתאות, לא נכלל בהוצאות)
ws3 = wb.create_sheet("דוחות סליקה גמא"); ws3.sheet_view.rightToLeft = True
header(ws3, ['#', 'תאריך', 'אסמכתא', 'ברוטו עסקאות (₪)', 'עמלות סליקה (₪)', 'קובץ'],
       [4, 12, 16, 18, 16, 30])
rr = 2
for i, g in enumerate(gama_dep, 1):
    vals = [i, g['date'], g['ref'], g['gross'], g['fee'], g['file']]
    for c, v in enumerate(vals, 1):
        cell = ws3.cell(rr, c, v); cell.font = Font(name=FN, size=10)
        cell.fill = ALT if rr % 2 == 0 else WHT; cell.border = BORD
        cell.alignment = Alignment('center', 'center')
        if c in (4, 5): cell.number_format = '#,##0.00'
    rr += 1
ws3.cell(rr + 1, 2, 'הערה: דוחות פיקדון/סליקה — הברוטו הוא הכנסה, לא הוצאה. '
                    'עמלת הסליקה (ההוצאה המוכרת) מתועדת בחשבוניות גמא בגיליון "הוצאות 2025".')
ws3.cell(rr + 1, 2).font = Font(name=FN, size=9, italic=True)

OUT = str(Path(__file__).parent / "הוצאות 25" / "סיכום_הוצאות_2025_מפורט.xlsx")
wb.save(OUT)
ok = sum(1 for r in rows if r['total'] is not None)
print(f"רשומות הוצאה: {len(rows)} | עם סכום: {ok} | מסומן לבדיקה: {sum(1 for r in rows if r['flag'])}")
print(f"דוחות סליקה גמא (אסמכתאות): {len(gama_dep)}")
print(f'סה"כ הוצאות: {sum(r["total"] or 0 for r in rows):,.2f} ₪')
print("נשמר:", OUT)
